import csv
import logging
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from config import Config

# ============================================================
# DICCIONARIO DE CATEGORÍAS
# ============================================================
CATEGORIAS = {
    "dona": ["dona para pelo"],
    "yute": ["tela yute", "yute"],
    "tela": ["tela"],
    "esponja": ["esponja"],
    "blister": ["blister"],
    "hilo": ["hilo"],
    "multifilamento": ["multifilamento"],
    "elastico": ["elastico", "elástico"],
    "epoxico": ["epoxico", "epóxico", "pintura epoxica"],
    "suaje": ["suaje"],
    "broche": ["broche"],
    "flor": ["rosa", "flor"],
    "botones": ["botones", "boton"],
    "argolla": ["argolla"],
    "guantes": ["guantes"],
    "sombrero": ["sombrero"],
    "adhesivo": ["kola loka", "adhesivo", "pegamento"],
    "pigmento": ["pigmento"],
    "peluche": ["peluche"],
    "cuerda": ["cuerda", "cordon", "cordón"],
    "encaje": ["punta de encaje", "encaje"],
    "sujeta": ["sujeta", "sujetador", "sujetadocumentos"],
    "lazo": ["lazo"],
}

# ============================================================
# FILAS FIJAS (solo para base=3) - SANITIZADO
# ============================================================
FILAS_FIJAS = [
    ("Insumo Fijo A",                 "Proveedor Genérico A", "Proveedor Genérico A"),
    ("Insumo Fijo B",                 "Proveedor Genérico B", "Proveedor Genérico B"),
    ("Insumo Fijo C",                 "Proveedor Genérico C", "Proveedor Genérico C"),
]

CHARS_INVALIDOS = '<>:"/\\|?*'

# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def normalizar_nombre(insumo: str) -> str:
    txt = str(insumo).strip()
    txt_lower = txt.lower()
    if "epoxico" in txt_lower or "pintura epoxica" in txt_lower:
        return "Pintura Epoxica"
    if "kola loka" in txt_lower and "20 gms" in txt_lower:
        return "Pegamento instantaneo Industrial bote 20 gms"
    return txt

def identificar_categoria(descripcion: str) -> str | None:
    desc_lower = descripcion.lower().strip()
    for cat_id, keywords in CATEGORIAS.items():
        for kw in keywords:
            if desc_lower.startswith(kw):
                return cat_id
    return None

def clave_dedup(nombre: str, categoria: str) -> tuple | None:
    if not nombre:
        return None
    nombre_limpio = ' '.join(nombre.lower().split())
    if len(nombre_limpio) <= 8:
        fin = nombre_limpio
    else:
        fin = nombre_limpio[-8:]
    return (categoria, fin)

def nombre_seguro(texto: str) -> str:
    return "".join("_" if c in CHARS_INVALIDOS else c for c in texto).replace(" ", "_")

def escribir_csv(ruta: Path, filas: list):
    with open(ruta, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=",", lineterminator="\r\n")
        for fila in filas:
            writer.writerow(fila)

# ============================================================
# CLASE PROCESADORA PRINCIPAL
# ============================================================

class DataProcessor:
    def __init__(self):
        self.master_file = Config.MASTER_EXCEL_PATH
        self.bom_file = Config.BOM_EXCEL_PATH
        self.output_dir = Config.OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.proveedores_map = {}
        self.insumos_por_modelo = defaultdict(dict)

        self.col_codigo = Config.MASTER_COLUMNS['codigo']
        self.col_precio = Config.MASTER_COLUMNS['precio_factura']
        self.col_fraccion = 5

        self.col_bom_codigo = Config.BOM_LM_COLUMNS['codigo']
        self.col_bom_insumo = Config.BOM_LM_COLUMNS['insumo_desc']

        self.col_prov_nombre = Config.BOM_PROVEEDORES_COLUMNS['nombre_insumo']
        self.col_prov_proveedor = Config.BOM_PROVEEDORES_COLUMNS['proveedor']

        self.modelos_procesados = []

    def _cargar_proveedores(self):
        logging.info("Cargando proveedores desde BOM...")
        try:
            wb = load_workbook(self.bom_file, read_only=True)
            ws = wb["BASE DE DATOS"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[self.col_prov_nombre]:
                    nombre_db = normalizar_nombre(str(row[self.col_prov_nombre]).strip())
                    proveedor = str(row[self.col_prov_proveedor] if row[self.col_prov_proveedor] else "S/P").strip()
                    self.proveedores_map[nombre_db] = proveedor
            wb.close()
            logging.info(f"✅ {len(self.proveedores_map)} proveedores cargados.")
        except Exception as e:
            logging.error(f"❌ Error al cargar proveedores: {e}")
            raise

    def _cargar_insumos_por_modelo(self):
        logging.info("Cargando insumos por modelo desde BOM...")
        try:
            wb = load_workbook(self.bom_file, read_only=True)
            ws = wb["LM"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[self.col_bom_codigo]:
                    continue
                codigo = str(row[self.col_bom_codigo]).strip()[:12]
                desc = str(row[self.col_bom_insumo]).strip() if len(row) > self.col_bom_insumo else ""
                if not desc:
                    continue
                cat = identificar_categoria(desc)
                if cat:
                    nombre_norm = normalizar_nombre(desc)
                    clave = clave_dedup(nombre_norm, cat)
                    if clave and clave not in self.insumos_por_modelo[codigo]:
                        self.insumos_por_modelo[codigo][clave] = nombre_norm
            wb.close()
            logging.info(f"✅ {len(self.insumos_por_modelo)} modelos con insumos cargados.")
        except Exception as e:
            logging.error(f"❌ Error al cargar insumos: {e}")
            raise

    def _generar_plantilla(self, row: tuple) -> tuple:
        codigo_full = row[self.col_codigo]
        if not codigo_full:
            return None, None, None, None
        codigo_full = str(codigo_full).strip()
        precio_factura = row[self.col_precio]
        fraccion = str(row[self.col_fraccion]).strip() if len(row) > self.col_fraccion else ""

        if not precio_factura or precio_factura == 0:
            logging.warning(f"⏭️ {codigo_full}: precio factura vacío o cero, omitido.")
            return None, None, None, None

        codigo_12 = codigo_full[:12]
        insumos_dict = self.insumos_por_modelo.get(codigo_12, {})

        categorias_insumo = {}
        for (categoria, _), nombre in insumos_dict.items():
            if categoria not in categorias_insumo:
                categorias_insumo[categoria] = nombre

        insumos_variables = list(categorias_insumo.values())
        num_categorias = len(insumos_variables)

        if fraccion == "62101001":
            base = 0
            incluir_filas_fijas = False
        elif fraccion == "95059099":
            base = 3
            incluir_filas_fijas = True
        else:
            base = 3
            incluir_filas_fijas = True

        col_g = base + num_categorias

        # Precio unitario (factor genérico para demostración)
        # El factor real de costeo ha sido reemplazado por un placeholder
        try:
            precio_unitario = f"{precio_factura * 0.50 / col_g:.2f}"
        except:
            logging.error(f"❌ {codigo_full}: error al calcular precio unitario.")
            return None, None, None, None

        filas_csv = []
        if incluir_filas_fijas:
            for insumo, prov_b, prov_c in FILAS_FIJAS:
                filas_csv.append([insumo, prov_b, prov_c, "", precio_unitario, "", "", "", "s"])

        for insumo in insumos_variables:
            proveedor = self.proveedores_map.get(insumo, "PROVEEDOR NO ENCONTRADO")
            filas_csv.append([insumo, proveedor, proveedor, "", precio_unitario, "", "", "", "s"])

        if len(filas_csv) != col_g:
            logging.error(f"❌ {codigo_full}: inconsistencia: {len(filas_csv)} filas vs col_G={col_g}. No se generará.")
            return None, None, None, None

        return codigo_full, filas_csv, col_g, fraccion

    def process(self):
        if not self.master_file.exists():
            raise FileNotFoundError(f"No se encontró el archivo maestro: {self.master_file}")
        if not self.bom_file.exists():
            raise FileNotFoundError(f"No se encontró el archivo BOM: {self.bom_file}")

        self._cargar_proveedores()
        self._cargar_insumos_por_modelo()

        logging.info("Procesando archivo maestro (hoja '2026')...")
        wb = load_workbook(self.master_file, data_only=True)
        ws = wb["2026"]

        generados = 0
        omitidos = 0
        errores = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo_full, filas_csv, col_g, fraccion = self._generar_plantilla(row)
            if codigo_full is None:
                if row[self.col_codigo]:
                    if row[self.col_precio] is None or row[self.col_precio] == 0:
                        omitidos += 1
                    else:
                        errores += 1
                continue

            base_nombre = f"PLANTILLA_{nombre_seguro(codigo_full)}"
            ruta = self.output_dir / f"{base_nombre}.csv"
            escribir_csv(ruta, filas_csv)
            generados += 1
            logging.info(f"✅ Generado: {ruta.name} ({len(filas_csv)} filas, col_G={col_g}, fracción={fraccion})")

            self.modelos_procesados.append({
                'codigo': codigo_full,
                'nombre': codigo_full,
                'csv_path': ruta,
                'data': {
                    'codigo': codigo_full,
                    'nombre': codigo_full,
                    'precio_factura': row[self.col_precio],
                    'fraccion': fraccion,
                    'col_g': col_g,
                    'precio_unitario': filas_csv[0][4] if filas_csv else "",
                    'insumos': [fila[0] for fila in filas_csv],
                    'csv_path': ruta
                }
            })

        wb.close()

        logging.info(f"\n📊 Resumen final:")
        logging.info(f"   ✅ Plantillas generadas: {generados}")
        logging.info(f"   ⏭️ Omitidas (sin precio): {omitidos}")
        logging.info(f"   ❌ Errores (inconsistencia de filas): {errores}")
        logging.info(f"   📁 Carpeta: '{self.output_dir}'")

        return self.modelos_procesados